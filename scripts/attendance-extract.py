#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from datetime import date, datetime, time

from openpyxl import load_workbook


def normalize_value(header, value):
    header = (header or "").strip().lower()
    if value is None:
        return ""
    if isinstance(value, datetime):
        if "time" in header and "date" not in header:
            return value.strftime("%H:%M")
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        if "minutes" in header:
            return value.hour * 60 + value.minute
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if "minutes" in header:
            return int(round(value))
        return value
    text = str(value).strip()
    return text


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: attendance-extract.py <xlsx path> [--output <json path>]")

    workbook_path = sys.argv[1]
    output_path = None
    if len(sys.argv) >= 4 and sys.argv[2] == "--output":
        output_path = Path(sys.argv[3])

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    sheets = []
    names = []
    seen_names = set()

    for ws in wb.worksheets:
        if ws.title.strip().lower() == "landing page":
            continue

        sheet_rows = []
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            record = {"rowNumber": row_number}
            for header, value in zip(headers, row):
                if not header:
                    continue
                key = str(header).strip()
                record[key] = normalize_value(key, value)
            sheet_rows.append(record)

            name = str(record.get("Name", "")).strip()
            if name:
                normalized = name.lower()
                if normalized not in seen_names:
                    seen_names.add(normalized)
                    names.append(name)

        sheets.append({
            "sheetName": ws.title,
            "rowCount": len(sheet_rows),
            "headers": [header for header in headers if header],
            "rows": sheet_rows
        })

    payload = {
        "sourceFile": workbook_path,
        "sheetCount": len(sheets),
        "uniqueNames": names,
        "sheets": sheets
    }

    if output_path:
        output_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        return

    json.dump(payload, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
